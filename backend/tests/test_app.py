"""End-to-end API tests in mock mode: admin bootstrap → create user →
user login → FBR settings → CSV upload → invoices submitted → receipt/QR →
soft deletes → deactivation."""

import os
import sys
from datetime import date, datetime, timedelta, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ["DATABASE_URL"] = "sqlite:///./test_fbr.db"
os.environ["ADMIN_EMAIL"] = "admin@example.com"
os.environ["ADMIN_PASSWORD"] = "admin123"

import jwt
import pytest
from fastapi.testclient import TestClient

if os.path.exists("test_fbr.db"):
    os.remove("test_fbr.db")

from app.config import get_settings
from app.main import app

client = TestClient(app)

CSV_OK = """pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,buyer_address,buyer_registration_type,product_description,hs_code,rate,uom,quantity,unit_price,sale_type,scenario_id
POS-1,2026-08-15,1234567,ABC Traders,Punjab,Lahore,Registered,Laptop,8471.3010,18%,"Numbers, pieces, units",2,150000,Goods at standard rate (default),SN001
POS-1,2026-08-15,1234567,ABC Traders,Punjab,Lahore,Registered,Mouse,8471.6020,18%,"Numbers, pieces, units",5,2500,Goods at standard rate (default),SN001
POS-2,2026-08-15,,Walk-in,Sindh,Karachi,Unregistered,Phone,8517.1219,18%,"Numbers, pieces, units",1,45000,Goods at standard rate (default),SN002
"""


@pytest.fixture(autouse=True, scope="module")
def cleanup():
    yield
    if os.path.exists("test_fbr.db"):
        os.remove("test_fbr.db")


def _login(email, password):
    resp = client.post("/api/auth/login", json={"email": email, "password": password})
    assert resp.status_code == 200, resp.text
    return {"Authorization": f"Bearer {resp.json()['token']}"}


@pytest.fixture(scope="module")
def admin_headers():
    return _login("admin@example.com", "admin123")


# Temp passwords are now server-generated (see test_create_user_...
# below) — this stashes one across two sequential tests that both need
# the same account's real temp password (create → later re-login).
_stash = {}

SHOP_REAL_PASSWORD = "Str0ng!Passw0rd99"


@pytest.fixture(scope="module")
def user_headers(admin_headers):
    resp = client.post(
        "/api/admin/users",
        json={"email": "shop@example.com", "full_name": "Shop Owner"},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    user_id = resp.json()["id"]
    shop_temp_password = resp.json()["temp_password"]
    assert len(shop_temp_password) == 6
    assert shop_temp_password.isalnum() and shop_temp_password == shop_temp_password.upper()
    # FBR settings are admin-managed only — configure them here so every
    # other test can assume a working seller profile is already in place.
    resp = client.put(
        f"/api/admin/users/{user_id}/fbr-settings",
        json={
            "fbr_env": "mock",
            "seller_ntn_cnic": "7654321",
            "seller_business_name": "Shop Owner Pvt Ltd",
            "seller_province": "Punjab",
            "seller_address": "Lahore",
            "default_scenario": "SN001",
        },
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text

    # Admin-issued passwords are temporary — first login must replace it
    # before the account can do real work (mirrors the actual product flow).
    login = client.post(
        "/api/auth/login",
        json={"email": "shop@example.com", "password": shop_temp_password},
    )
    assert login.status_code == 200, login.text
    assert login.json()["must_change_password"] is True
    temp_headers = {"Authorization": f"Bearer {login.json()['token']}"}
    resp = client.post(
        "/api/auth/set-password",
        json={
            "new_password": SHOP_REAL_PASSWORD,
            "confirm_password": SHOP_REAL_PASSWORD,
        },
        headers=temp_headers,
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["must_change_password"] is False

    return _login("shop@example.com", SHOP_REAL_PASSWORD)


def test_password_strength_scoring():
    from app.auth import password_strength

    # The only real requirement is length >= 8 — no complexity scoring.
    assert password_strength("short1!")["ok"] is False  # 7 chars
    assert password_strength("abcdefgh")["ok"] is True  # 8 chars, no complexity needed
    assert password_strength("alllowercase")["ok"] is True
    assert password_strength("Str0ng!Passw0rd")["ok"] is True
    assert password_strength("short1!")["length"] == 7


def test_must_change_password_gates_real_actions(admin_headers):
    # A fresh admin-created user, never through the fixture's set-password step.
    resp = client.post(
        "/api/admin/users",
        json={"email": "temp@example.com", "full_name": "Temp User"},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    user_id = resp.json()["id"]
    assert resp.json()["must_change_password"] is True
    temp_password = resp.json()["temp_password"]

    login = client.post(
        "/api/auth/login", json={"email": "temp@example.com", "password": temp_password}
    )
    assert login.status_code == 200
    assert login.json()["must_change_password"] is True
    temp_headers = {"Authorization": f"Bearer {login.json()['token']}"}

    # Real work is blocked until the password is replaced.
    resp = client.post(
        "/api/uploads",
        files={"file": ("x.csv", "a,b\n1,2\n", "text/csv")},
        headers=temp_headers,
    )
    assert resp.status_code == 403
    assert "set your own password" in resp.json()["detail"]

    # Weak password rejected server-side, regardless of the client meter.
    resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "weak", "confirm_password": "weak"},
        headers=temp_headers,
    )
    assert resp.status_code == 400

    # Mismatched confirmation rejected.
    resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "Str0ng!Passw0rd", "confirm_password": "Different1!"},
        headers=temp_headers,
    )
    assert resp.status_code == 400

    # Strong, matching password succeeds.
    resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "Str0ng!Passw0rd", "confirm_password": "Str0ng!Passw0rd"},
        headers=temp_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["must_change_password"] is False
    fresh_headers = {"Authorization": f"Bearer {resp.json()['token']}"}

    # The OLD token (the one that made this very set-password call) is now
    # dead — token_version bumped, so it can't be reused to change the
    # password again or do anything else. Must use the fresh token above.
    resp = client.get("/api/auth/me", headers=temp_headers)
    assert resp.status_code == 401
    resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "AnotherStrong1!", "confirm_password": "AnotherStrong1!"},
        headers=temp_headers,
    )
    assert resp.status_code == 401

    # Uploads no longer blocked by the password gate (fails for a different,
    # expected reason — bad CSV columns — proving the 403 gate is gone).
    resp = client.post(
        "/api/uploads",
        files={"file": ("x.csv", "a,b\n1,2\n", "text/csv")},
        headers=fresh_headers,
    )
    assert resp.status_code == 201
    assert resp.json()["status"] == "failed"

    # Re-login with the new password succeeds and no longer flags the account.
    relogin = client.post(
        "/api/auth/login", json={"email": "temp@example.com", "password": "Str0ng!Passw0rd"}
    )
    assert relogin.status_code == 200
    assert relogin.json()["must_change_password"] is False

    # Old temp password no longer works.
    old_login = client.post(
        "/api/auth/login", json={"email": "temp@example.com", "password": temp_password}
    )
    assert old_login.status_code == 401

    client.delete(f"/api/admin/users/{user_id}", headers=admin_headers)


def test_admin_with_temp_password_blocked_from_admin_routes(admin_headers):
    # A freshly created SECOND admin, still on their temp password, must not
    # get admin power for free — otherwise the whole gate is pointless for
    # admin accounts specifically.
    resp = client.post(
        "/api/admin/users",
        json={"email": "newadmin@example.com", "full_name": "New Admin", "role": "admin"},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    new_admin_id = resp.json()["id"]
    new_admin_temp_password = resp.json()["temp_password"]

    login = client.post(
        "/api/auth/login",
        json={"email": "newadmin@example.com", "password": new_admin_temp_password},
    )
    assert login.status_code == 200
    assert login.json()["must_change_password"] is True
    temp_admin_headers = {"Authorization": f"Bearer {login.json()['token']}"}

    # Admin-role check passes (it IS an admin) but the password gate blocks
    # every admin action until they set their own password.
    assert client.get("/api/admin/users", headers=temp_admin_headers).status_code == 403
    assert (
        client.post(
            "/api/admin/users",
            json={"email": "x@example.com", "full_name": "X"},
            headers=temp_admin_headers,
        ).status_code
        == 403
    )

    resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "Str0ng!AdminPass", "confirm_password": "Str0ng!AdminPass"},
        headers=temp_admin_headers,
    )
    assert resp.status_code == 200
    fresh_admin_headers = {"Authorization": f"Bearer {resp.json()['token']}"}

    # Now admin routes work normally.
    assert client.get("/api/admin/users", headers=fresh_admin_headers).status_code == 200

    client.delete(f"/api/admin/users/{new_admin_id}", headers=admin_headers)


def test_password_strength_endpoint(user_headers):
    resp = client.post("/api/auth/password-strength", json={"password": "Str0ng!Passw0rd"})
    assert resp.status_code == 200
    assert resp.json()["ok"] is True

    resp = client.post("/api/auth/password-strength", json={"password": "weak"})
    assert resp.json()["ok"] is False


def test_call_handles_malformed_json_response_without_crashing(monkeypatch):
    # Regression test: FBR returning 200 OK with a non-JSON/malformed body
    # (a real incident hit live) must not let json.JSONDecodeError propagate
    # uncaught past _call() — it should retry, then raise a clean FBRError.
    import httpx as httpx_module

    from app.fbr import client
    from app.models import FbrSettings

    calls = {"n": 0}

    class FakeResponse:
        status_code = 200
        text = '{"validationResponse": {"statusCode": "00" "error": "}'  # malformed

        def raise_for_status(self):
            pass

        def json(self):
            import json as json_module

            return json_module.loads(self.text)  # raises JSONDecodeError

    def fake_post(url, headers, json, timeout):
        calls["n"] += 1
        return FakeResponse()

    monkeypatch.setattr(httpx_module, "post", fake_post)
    monkeypatch.setattr(client.time, "sleep", lambda _: None)  # skip real backoff

    fbr = FbrSettings(fbr_env="sandbox", fbr_token="test-token")
    with pytest.raises(client.FBRError, match="failed to parse as JSON"):
        client._call("https://example.invalid/x", {}, fbr)

    # Retried the configured number of times (default retries=2 -> 3 attempts).
    assert calls["n"] == 3


def test_is_valid_checks_item_level_statuses():
    # Spec v1.12 §4.1.5: outer statusCode can be "00" while items fail.
    from app.fbr.client import is_valid

    assert is_valid(
        {"validationResponse": {"statusCode": "00", "invoiceStatuses": [{"statusCode": "00"}]}}
    )
    assert not is_valid(
        {
            "validationResponse": {
                "statusCode": "00",
                "status": "invalid",
                "invoiceStatuses": [{"statusCode": "00"}, {"statusCode": "01"}],
            }
        }
    )
    assert not is_valid({"validationResponse": {"statusCode": "01"}})
    # invoiceStatuses may be null (spec §4.1.4)
    assert is_valid({"validationResponse": {"statusCode": "00", "invoiceStatuses": None}})


def test_applicable_scenarios_matrix():
    from app.fbr.scenarios import applicable_scenarios

    assert applicable_scenarios("Manufacturer", "Steel") == ["SN003", "SN004", "SN011"]
    assert "SN025" not in applicable_scenarios("Manufacturer", "Pharmaceuticals")
    assert "SN025" in applicable_scenarios("Importer", "Pharmaceuticals")
    assert applicable_scenarios("Service Provider", "Services") == ["SN018", "SN019"]
    retailer_fmcg = applicable_scenarios("Retailer", "FMCG")
    assert retailer_fmcg == ["SN008", "SN026", "SN027", "SN028"]


def test_error_code_guidance_appended():
    from app.fbr.client import error_text

    text = error_text(
        {
            "validationResponse": {
                "statusCode": "01",
                "errorCode": "0019",
                "error": "HS Code is either not provided or invalid",
            }
        }
    )
    # Original FBR message preserved, plus PRAL's published remediation.
    assert "HS Code is either not provided or invalid" in text
    assert "reference API" in text

    # Unrecognized code: message passes through unchanged, no crash.
    text2 = error_text(
        {"validationResponse": {"statusCode": "01", "errorCode": "9999", "error": "Some new error"}}
    )
    assert text2 == "Some new error"


def test_scenario_sample_data_covers_all_28():
    from app.fbr.scenario_samples import SCENARIO_SAMPLE_ITEMS

    assert len(SCENARIO_SAMPLE_ITEMS) == 28
    for code in (f"SN{n:03d}" for n in range(1, 29)):
        item = SCENARIO_SAMPLE_ITEMS[code]
        assert item["hsCode"]
        assert item["saleType"]
        assert item["quantity"] > 0
        # 3rd Schedule / retail-reduced-rate goods (SN008, SN027, SN028) are
        # priced via fixedNotifiedValueOrRetailPrice instead —
        # valueSalesExcludingST is legitimately 0 for those, not missing data.
        assert item["valueSalesExcludingST"] > 0 or item["fixedNotifiedValueOrRetailPrice"] > 0


def test_third_schedule_scenarios_price_off_fixed_value():
    from app.fbr.scenario_samples import SCENARIO_SAMPLE_ITEMS

    for code in ("SN008", "SN027", "SN028"):
        item = SCENARIO_SAMPLE_ITEMS[code]
        assert item["valueSalesExcludingST"] == 0
        assert item["fixedNotifiedValueOrRetailPrice"] > 0

    # SN028's official sample also requires an SRO schedule reference.
    sn028 = SCENARIO_SAMPLE_ITEMS["SN028"]
    assert sn028["sroScheduleNo"]
    assert sn028["sroItemSerialNo"]




def test_scenario_aware_csv_template(user_headers):
    resp = client.get("/api/uploads/template?scenario=SN015", headers=user_headers)
    assert resp.status_code == 200
    rows = resp.text.strip().splitlines()
    assert len(rows) == 2  # header + one official-sample row
    assert "SN015" in rows[1]

    # Unknown scenario -> 404, not a silent empty file.
    resp = client.get("/api/uploads/template?scenario=SN999", headers=user_headers)
    assert resp.status_code == 404


def test_scenario_catalog(user_headers):
    resp = client.get("/api/uploads/scenario-catalog", headers=user_headers)
    assert resp.status_code == 200
    codes = {s["code"] for s in resp.json()}
    assert codes == {f"SN{n:03d}" for n in range(1, 29)}


def test_every_scenario_csv_template_parses_cleanly(user_headers):
    # Each of the 28 official-sample templates must itself pass our own CSV
    # pipeline — otherwise "download the sample for this scenario" would
    # hand the customer a file our own uploader rejects. Parse-only (not a
    # live upload) so this doesn't disturb invoice counts other tests assert.
    from app.services.csv_processor import _parse_rows

    for n in range(1, 29):
        code = f"SN{n:03d}"
        resp = client.get(f"/api/uploads/template?scenario={code}", headers=user_headers)
        assert resp.status_code == 200, code
        rows = _parse_rows(resp.text)  # raises CsvError if malformed
        assert len(rows) == 1 and rows[0]["scenario_id"] == code


def test_login_rejects_bad_password():
    resp = client.post(
        "/api/auth/login",
        json={"email": "admin@example.com", "password": "wrong"},
    )
    assert resp.status_code == 401


def test_auth_required():
    assert client.get("/api/invoices").status_code == 401
    assert client.get("/api/admin/users").status_code == 401


def test_admin_only(user_headers):
    assert client.get("/api/admin/users", headers=user_headers).status_code == 403


def test_fbr_settings_read_only_for_user(user_headers):
    # GET still works — users can view the settings an admin configured.
    data = client.get("/api/settings/fbr", headers=user_headers).json()
    assert data["seller_business_name"] == "Shop Owner Pvt Ltd"
    assert data["has_sandbox_token"] is False
    assert data["has_production_token"] is False
    assert data["can_submit_production"] is False

    # PUT is retired for regular users — 403, not a silent no-op.
    resp = client.put(
        "/api/settings/fbr",
        json={"fbr_env": "sandbox", "seller_business_name": "Hacked Name"},
        headers=user_headers,
    )
    assert resp.status_code == 403

    # An unauthenticated caller gets 401 (auth checked first), not the
    # admin-specific 403 message, which would be misleading with no token.
    resp = client.put(
        "/api/settings/fbr", json={"fbr_env": "sandbox"}
    )
    assert resp.status_code == 401

    # Confirm nothing changed.
    data2 = client.get("/api/settings/fbr", headers=user_headers).json()
    assert data2["seller_business_name"] == "Shop Owner Pvt Ltd"
    assert data2["fbr_env"] == "mock"


def test_admin_manages_user_fbr_settings(admin_headers, user_headers):
    users = client.get("/api/admin/users", headers=admin_headers).json()
    shop = next(u for u in users if u["email"] == "shop@example.com")
    ORIGINAL_PROFILE = {
        "fbr_env": "mock",
        "seller_ntn_cnic": "7654321",
        "seller_business_name": "Shop Owner Pvt Ltd",
        "seller_province": "Punjab",
        "seller_address": "Lahore",
        "default_scenario": "SN001",
    }

    try:
        # A non-admin (even for their own account) cannot reach the admin route.
        resp = client.put(
            f"/api/admin/users/{shop['id']}/fbr-settings",
            json={"fbr_env": "sandbox"},
            headers=user_headers,
        )
        assert resp.status_code == 403

        resp = client.get(
            f"/api/admin/users/{shop['id']}/fbr-settings", headers=user_headers
        )
        assert resp.status_code == 403

        # Admin can read and update it.
        resp = client.get(
            f"/api/admin/users/{shop['id']}/fbr-settings", headers=admin_headers
        )
        assert resp.status_code == 200
        assert resp.json()["seller_business_name"] == "Shop Owner Pvt Ltd"

        resp = client.put(
            f"/api/admin/users/{shop['id']}/fbr-settings",
            json={
                "fbr_env": "mock",
                "seller_ntn_cnic": "9999999",
                "seller_ntn": "1234567",
                "seller_business_name": "Renamed By Admin",
                "seller_province": "Sindh",
                "seller_address": "Karachi",
                "default_scenario": "SN002",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200
        assert resp.json()["seller_business_name"] == "Renamed By Admin"
        assert resp.json()["seller_ntn"] == "1234567"

        # The user sees the admin's change but still cannot edit it themselves.
        seen = client.get("/api/settings/fbr", headers=user_headers).json()
        assert seen["seller_business_name"] == "Renamed By Admin"
        assert seen["seller_ntn"] == "1234567"
        # seller_ntn is display-only — never sent to FBR as sellerNTNCNIC.
        assert seen["seller_ntn_cnic"] == "9999999"
    finally:
        # Always restore the shared profile, even if an assertion above
        # failed — later tests (e.g. test_csv_upload_and_receipts) hard-code
        # this exact seller profile, and a skipped restore would corrupt
        # module-scoped state and mask the real failure in an unrelated test.
        resp = client.put(
            f"/api/admin/users/{shop['id']}/fbr-settings",
            json=ORIGINAL_PROFILE,
            headers=admin_headers,
        )
        assert resp.status_code == 200


def test_csv_template(user_headers):
    resp = client.get("/api/uploads/template", headers=user_headers)
    assert resp.status_code == 200
    assert "pos_invoice_no" in resp.text
    assert "scenario_id" in resp.text  # the shared fixture's user is in mock mode

    # Unauthenticated download is no longer allowed — the response has to
    # be shaped by the requesting user's own FBR environment.
    assert client.get("/api/uploads/template").status_code == 401


def test_production_account_template_omits_scenario_id(admin_headers, user_headers):
    users = client.get("/api/admin/users", headers=admin_headers).json()
    shop = next(u for u in users if u["email"] == "shop@example.com")
    ORIGINAL_PROFILE = {
        "fbr_env": "mock",
        "seller_ntn_cnic": "7654321",
        "seller_business_name": "Shop Owner Pvt Ltd",
        "seller_province": "Punjab",
        "seller_address": "Lahore",
        "default_scenario": "SN001",
    }
    try:
        client.put(
            f"/api/admin/users/{shop['id']}/fbr-settings",
            json={**ORIGINAL_PROFILE, "fbr_env": "production"},
            headers=admin_headers,
        )
        resp = client.get("/api/uploads/template", headers=user_headers)
        assert resp.status_code == 200
        header_row = resp.text.splitlines()[0]
        assert "scenario_id" not in header_row.split(",")
        assert "pos_invoice_no" in header_row

        # A scenario-specific sandbox download still carries scenario_id
        # regardless of env — the frontend just never surfaces it to a
        # production account in the first place.
        scenario_resp = client.get(
            "/api/uploads/template?scenario=SN001", headers=user_headers
        )
        assert "scenario_id" in scenario_resp.text.splitlines()[0]
    finally:
        client.put(
            f"/api/admin/users/{shop['id']}/fbr-settings",
            json=ORIGINAL_PROFILE,
            headers=admin_headers,
        )


def test_csv_upload_and_receipts(user_headers):
    resp = client.post(
        "/api/uploads",
        files={"file": ("sales.csv", CSV_OK, "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    upload = resp.json()
    assert upload["status"] == "completed"
    assert upload["total_rows"] == 3
    assert upload["invoices_created"] == 2  # POS-1 (2 items) + POS-2 (1 item)
    assert upload["invoices_submitted"] == 2

    invoices = client.get("/api/invoices", headers=user_headers).json()
    assert len(invoices) == 2
    assert all(i["status"] == "submitted" for i in invoices)
    assert all(i["fbr_invoice_number"].startswith("MOCK") for i in invoices)

    # POS-1: 2×150000 + 5×2500 = 312500 excl; 18% tax = 56250
    pos1 = next(i for i in invoices if i["pos_invoice_no"] == "POS-1")
    assert pos1["total_excl"] == 312500.0
    assert pos1["total_tax"] == 56250.0

    detail = client.get(f"/api/invoices/{pos1['id']}", headers=user_headers).json()
    assert detail["qr"].startswith("data:image/png;base64")
    assert detail["payload"]["sellerNTNCNIC"] == "7654321"
    assert detail["payload"]["scenarioId"] == "SN001"
    assert len(detail["items"]) == 2
    assert detail["fbr_response"]["validationResponse"]["statusCode"] == "00"
    # Spec v1.12: extraTax is numeric; mock returns per-item statuses
    assert isinstance(detail["payload"]["items"][0]["extraTax"], (int, float))
    item_statuses = detail["fbr_response"]["validationResponse"]["invoiceStatuses"]
    assert len(item_statuses) == 2
    assert all(s["statusCode"] == "00" for s in item_statuses)


def test_csv_upload_bad_format(user_headers):
    resp = client.post(
        "/api/uploads",
        files={"file": ("bad.csv", "foo,bar\n1,2\n", "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201
    assert resp.json()["status"] == "failed"
    assert "Missing required columns" in resp.json()["error"]


def test_users_cannot_delete_uploads(user_headers):
    # An upload can only be removed by an admin — the user-facing DELETE
    # route doesn't exist (so this is a no-op 404, nothing is removed).
    # Deleting a *test* invoice is allowed — see
    # test_user_can_delete_test_invoice_only.
    up_id = client.get("/api/uploads", headers=user_headers).json()[0]["id"]
    assert client.delete(
        f"/api/uploads/{up_id}", headers=user_headers
    ).status_code in (404, 405)


def test_admin_stats(admin_headers):
    stats = client.get("/api/admin/stats", headers=admin_headers).json()
    assert stats["total_users"] == 2
    assert stats["total_invoices"] == 2
    assert stats["submitted_invoices"] == 2

def _make_account(admin_headers, email, **fbr_overrides):
    """Create a throwaway user with a real password + FBR settings, return
    its bearer headers and id. Keeps env-specific tests off the shared
    shop@example.com account whose invoice counts other tests assert on."""
    resp = client.post(
        "/api/admin/users",
        json={"email": email, "full_name": email.split("@")[0]},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    uid = resp.json()["id"]
    temp = resp.json()["temp_password"]
    settings = {
        "fbr_env": "mock",
        "seller_ntn_cnic": "7654321",
        "seller_business_name": "Promo Pvt Ltd",
        "seller_province": "Sindh",
        "seller_address": "Karachi",
        "default_scenario": "SN001",
        **fbr_overrides,
    }
    assert (
        client.put(
            f"/api/admin/users/{uid}/fbr-settings", json=settings, headers=admin_headers
        ).status_code
        == 200
    )
    login = client.post(
        "/api/auth/login", json={"email": email, "password": temp}
    )
    th = {"Authorization": f"Bearer {login.json()['token']}"}
    done = client.post(
        "/api/auth/set-password",
        json={"new_password": "Str0ng!Passw0rd99", "confirm_password": "Str0ng!Passw0rd99"},
        headers=th,
    )
    return {"Authorization": f"Bearer {done.json()['token']}"}, uid


_ENV_CSV = (
    "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
    "buyer_address,buyer_registration_type,product_description,hs_code,"
    "rate,uom,quantity,unit_price,sale_type,scenario_id\n"
    "POS-ENV-1,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
    "Widget,0101.2100,18%,\"Numbers, pieces, units\",1,1000,"
    "Goods at standard rate (default),SN002\n"
)


def test_admin_manages_seller_strns(admin_headers):
    headers, uid = _make_account(admin_headers, "strns@example.com")
    base = {
        "fbr_env": "mock",
        "seller_ntn_cnic": "7654321",
        "seller_business_name": "Strn Pvt Ltd",
        "seller_province": "Sindh",
        "seller_address": "Karachi",
        "default_scenario": "SN001",
    }

    def put(extra):
        r = client.put(
            f"/api/admin/users/{uid}/fbr-settings",
            json={**base, **extra},
            headers=admin_headers,
        )
        assert r.status_code == 200, r.text
        return r.json()

    # Several entries round-trip; each half is trimmed and rows missing
    # either half are dropped.
    out = put(
        {
            "strns": [
                {"business_name": "  Alpha Traders ", "strn": " 111 "},
                {"business_name": "Beta Foods", "strn": "222"},
                {"business_name": "", "strn": "333"},
                {"business_name": "No Number", "strn": ""},
            ]
        }
    )
    assert out["strns"] == [
        {"business_name": "Alpha Traders", "strn": "111"},
        {"business_name": "Beta Foods", "strn": "222"},
    ]

    # The user sees them, read-only.
    assert (
        client.get("/api/settings/fbr", headers=headers).json()["strns"]
        == out["strns"]
    )

    # Omitting the key leaves the saved list untouched; [] clears it.
    assert len(put({})["strns"]) == 2
    assert put({"strns": []})["strns"] == []


def test_invoice_receipt_includes_seller_strns(admin_headers):
    headers, _ = _make_account(
        admin_headers,
        "strnreceipt@example.com",
        strns=[
            {"business_name": "Alpha", "strn": "111"},
            {"business_name": "Beta", "strn": "222"},
        ],
    )
    up = client.post(
        "/api/uploads",
        files={"file": ("s.csv", _ENV_CSV, "text/csv")},
        data={"target": "sandbox"},
        headers=headers,
    ).json()
    inv_id = client.get(
        f"/api/invoices?upload_id={up['id']}", headers=headers
    ).json()[0]["id"]
    detail = client.get(f"/api/invoices/{inv_id}", headers=headers).json()
    assert detail["seller"]["strns"] == [
        {"business_name": "Alpha", "strn": "111"},
        {"business_name": "Beta", "strn": "222"},
    ]


def test_upload_target_env_is_stamped_and_history_splits(admin_headers):
    headers, uid = _make_account(
        admin_headers, "envsplit@example.com", can_submit_production=True
    )

    sbx = client.post(
        "/api/uploads",
        files={"file": ("s.csv", _ENV_CSV, "text/csv")},
        data={"target": "sandbox"},
        headers=headers,
    ).json()
    prod = client.post(
        "/api/uploads",
        files={"file": ("p.csv", _ENV_CSV, "text/csv")},
        data={"target": "production"},
        headers=headers,
    ).json()
    assert sbx["fbr_env"] == "sandbox" and sbx["invoices_submitted"] == 1
    assert prod["fbr_env"] == "production" and prod["invoices_submitted"] == 1

    # Submission History splits by ?fbr_env=
    only_sbx = client.get("/api/uploads?fbr_env=sandbox", headers=headers).json()
    only_prod = client.get("/api/uploads?fbr_env=production", headers=headers).json()
    assert [u["id"] for u in only_sbx] == [sbx["id"]]
    assert [u["id"] for u in only_prod] == [prod["id"]]

    # Invoices History splits the same way, and each invoice carries its env.
    inv_sbx = client.get("/api/invoices?fbr_env=sandbox", headers=headers).json()
    inv_prod = client.get("/api/invoices?fbr_env=production", headers=headers).json()
    assert len(inv_sbx) == 1 and inv_sbx[0]["fbr_env"] == "sandbox"
    assert len(inv_prod) == 1 and inv_prod[0]["fbr_env"] == "production"
    # A mock account simulates production too — invoice number still MOCK…
    assert inv_prod[0]["fbr_invoice_number"].startswith("MOCK")

    # bad env value → 400
    assert client.get("/api/invoices?fbr_env=bogus", headers=headers).status_code == 400


def test_production_upload_blocked_without_capability(admin_headers):
    headers, _ = _make_account(admin_headers, "nocap@example.com")  # can_submit_production defaults False
    resp = client.post(
        "/api/uploads",
        files={"file": ("p.csv", _ENV_CSV, "text/csv")},
        data={"target": "production"},
        headers=headers,
    )
    assert resp.status_code == 403


def test_promote_invoice_to_production(admin_headers):
    headers, _ = _make_account(
        admin_headers, "promote@example.com", can_submit_production=True
    )
    up = client.post(
        "/api/uploads",
        files={"file": ("s.csv", _ENV_CSV, "text/csv")},
        data={"target": "sandbox"},
        headers=headers,
    ).json()
    inv = client.get(
        f"/api/invoices?upload_id={up['id']}", headers=headers
    ).json()[0]
    assert inv["fbr_env"] == "sandbox" and inv["status"] == "submitted"
    sandbox_number = inv["fbr_invoice_number"]

    promoted = client.post(
        f"/api/invoices/{inv['id']}/promote", headers=headers
    )
    assert promoted.status_code == 200, promoted.text
    body = promoted.json()
    assert body["fbr_env"] == "production"
    assert body["status"] == "submitted"
    assert body["fbr_invoice_number"] != sandbox_number

    # It has moved out of sandbox history and into production history.
    assert client.get("/api/invoices?fbr_env=sandbox", headers=headers).json() == []
    prod_hist = client.get("/api/invoices?fbr_env=production", headers=headers).json()
    assert [i["id"] for i in prod_hist] == [inv["id"]]

    # Promoting again → 400 (already production).
    assert (
        client.post(f"/api/invoices/{inv['id']}/promote", headers=headers).status_code
        == 400
    )


def test_promote_whole_upload_to_production(admin_headers):
    headers, _ = _make_account(
        admin_headers, "promotebatch@example.com", can_submit_production=True
    )
    two_row = _ENV_CSV + (
        "POS-ENV-2,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Widget2,0101.2100,18%,\"Numbers, pieces, units\",3,500,"
        "Goods at standard rate (default),SN002\n"
    )
    up = client.post(
        "/api/uploads",
        files={"file": ("batch.csv", two_row, "text/csv")},
        data={"target": "sandbox"},
        headers=headers,
    ).json()
    assert up["fbr_env"] == "sandbox" and up["invoices_submitted"] == 2

    resp = client.post(f"/api/uploads/{up['id']}/promote", headers=headers)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["fbr_env"] == "production"
    assert body["invoices_submitted"] == 2 and body["invoices_failed"] == 0

    # Both invoices moved to production history; batch left the test history.
    assert client.get("/api/invoices?fbr_env=sandbox", headers=headers).json() == []
    assert len(client.get("/api/invoices?fbr_env=production", headers=headers).json()) == 2
    assert client.get("/api/uploads?fbr_env=sandbox", headers=headers).json() == []
    assert [u["id"] for u in client.get(
        "/api/uploads?fbr_env=production", headers=headers
    ).json()] == [up["id"]]

    # Re-promoting the batch → 400.
    assert (
        client.post(f"/api/uploads/{up['id']}/promote", headers=headers).status_code == 400
    )

    # A non-capable account cannot promote a batch.
    plain, _ = _make_account(admin_headers, "promotebatchno@example.com")
    up2 = client.post(
        "/api/uploads",
        files={"file": ("s.csv", _ENV_CSV, "text/csv")},
        data={"target": "sandbox"},
        headers=plain,
    ).json()
    assert (
        client.post(f"/api/uploads/{up2['id']}/promote", headers=plain).status_code == 403
    )


def test_promoting_invoices_individually_flips_the_parent_batch(admin_headers):
    # Reported bug: promoting every invoice in a batch one by one left its
    # Submission History row stuck on "Test" — only the batch-level promote
    # flipped the upload. Now the upload rolls up from its invoices.
    headers, _ = _make_account(
        admin_headers, "promoteindiv@example.com", can_submit_production=True
    )
    two_row = _ENV_CSV + (
        "POS-ENV-2,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Widget2,0101.2100,18%,\"Numbers, pieces, units\",3,500,"
        "Goods at standard rate (default),SN002\n"
    )
    up = client.post(
        "/api/uploads",
        files={"file": ("batch.csv", two_row, "text/csv")},
        data={"target": "sandbox"},
        headers=headers,
    ).json()
    assert up["fbr_env"] == "sandbox"
    inv_ids = [
        i["id"]
        for i in client.get(
            f"/api/invoices?upload_id={up['id']}", headers=headers
        ).json()
    ]
    assert len(inv_ids) == 2

    def batch_env():
        rows = client.get("/api/uploads?fbr_env=all", headers=headers).json()
        return next(u for u in rows if u["id"] == up["id"])["fbr_env"]

    # First of two promoted — a test invoice remains, batch stays "Test".
    assert (
        client.post(
            f"/api/invoices/{inv_ids[0]}/promote", headers=headers
        ).status_code
        == 200
    )
    assert batch_env() == "sandbox"

    # Second (last) promoted — nothing test left, batch flips to Live.
    assert (
        client.post(
            f"/api/invoices/{inv_ids[1]}/promote", headers=headers
        ).status_code
        == 200
    )
    assert batch_env() == "production"
    assert client.get("/api/uploads?fbr_env=sandbox", headers=headers).json() == []

    # The now fully-live batch refuses a batch promote.
    assert (
        client.post(
            f"/api/uploads/{up['id']}/promote", headers=headers
        ).status_code
        == 400
    )


def test_promote_guards(admin_headers):
    # No capability → 403.
    plain, _ = _make_account(admin_headers, "promoteno@example.com")
    up = client.post(
        "/api/uploads",
        files={"file": ("s.csv", _ENV_CSV, "text/csv")},
        data={"target": "sandbox"},
        headers=plain,
    ).json()
    inv_id = client.get(
        f"/api/invoices?upload_id={up['id']}", headers=plain
    ).json()[0]["id"]
    assert client.post(f"/api/invoices/{inv_id}/promote", headers=plain).status_code == 403

    # Capable, but the invoice never submitted cleanly → 400.
    cap, _ = _make_account(
        admin_headers, "promotebad@example.com", can_submit_production=True
    )
    bad_csv = _ENV_CSV.replace(
        "\"Numbers, pieces, units\",1,1000", "\"Numbers, pieces, units\",0,1000"
    )  # quantity 0 trips the mock validator
    up2 = client.post(
        "/api/uploads",
        files={"file": ("bad.csv", bad_csv, "text/csv")},
        data={"target": "sandbox"},
        headers=cap,
    ).json()
    bad_inv = client.get(
        f"/api/invoices?upload_id={up2['id']}", headers=cap
    ).json()[0]
    assert bad_inv["status"] == "failed"
    assert (
        client.post(f"/api/invoices/{bad_inv['id']}/promote", headers=cap).status_code
        == 400
    )


def test_stats_split_by_test_vs_live(admin_headers):
    headers, _ = _make_account(
        admin_headers, "envstats@example.com", can_submit_production=True
    )
    # 1 test batch (2 invoices), 1 live batch (1 invoice).
    two = _ENV_CSV + (
        "POS-ENV-2,2026-08-17,,W,Sindh,Karachi,Unregistered,Y,0101.2100,18%,"
        "\"Numbers, pieces, units\",1,500,Goods at standard rate (default),SN002\n"
    )
    client.post(
        "/api/uploads",
        files={"file": ("t.csv", two, "text/csv")},
        data={"target": "sandbox"},
        headers=headers,
    )
    client.post(
        "/api/uploads",
        files={"file": ("l.csv", _ENV_CSV, "text/csv")},
        data={"target": "production"},
        headers=headers,
    )

    all_s = client.get("/api/stats", headers=headers).json()
    test_s = client.get("/api/stats?fbr_env=test", headers=headers).json()
    live_s = client.get("/api/stats?fbr_env=live", headers=headers).json()

    assert all_s["total_invoices"] == 3 and all_s["total_uploads"] == 2
    assert test_s["total_invoices"] == 2 and test_s["total_uploads"] == 1
    assert live_s["total_invoices"] == 1 and live_s["total_uploads"] == 1
    assert test_s["submitted_invoices"] + live_s["submitted_invoices"] == all_s["submitted_invoices"]
    # "sandbox" is an alias for test (covers mock too).
    assert client.get("/api/stats?fbr_env=sandbox", headers=headers).json()["total_invoices"] == 2
    assert client.get("/api/stats?fbr_env=bogus", headers=headers).status_code == 400


def test_user_isolation(admin_headers):
    resp = client.post(
        "/api/admin/users",
        json={"email": "other@example.com", "full_name": "Other"},
        headers=admin_headers,
    )
    _stash["other_temp_password"] = resp.json()["temp_password"]
    other = _login("other@example.com", _stash["other_temp_password"])
    assert client.get("/api/invoices", headers=other).json() == []


def test_deactivate_blocks_login(admin_headers):
    users = client.get("/api/admin/users", headers=admin_headers).json()
    other = next(u for u in users if u["email"] == "other@example.com")
    resp = client.patch(
        f"/api/admin/users/{other['id']}",
        json={"is_active": False},
        headers=admin_headers,
    )
    assert resp.status_code == 200
    resp = client.post(
        "/api/auth/login",
        json={"email": "other@example.com", "password": _stash["other_temp_password"]},
    )
    assert resp.status_code == 403


def test_soft_delete_user(admin_headers):
    users = client.get("/api/admin/users", headers=admin_headers).json()
    other = next(u for u in users if u["email"] == "other@example.com")
    assert client.delete(
        f"/api/admin/users/{other['id']}", headers=admin_headers
    ).json() == {"ok": True}
    users = client.get("/api/admin/users", headers=admin_headers).json()
    assert all(u["email"] != "other@example.com" for u in users)


def test_recreate_soft_deleted_email_restores_account(admin_headers):
    # The email of a soft-deleted user must be reusable: creating it again
    # restores the account (same row — users.email is UNIQUE in the DB).
    resp = client.post(
        "/api/admin/users",
        json={"email": "other@example.com", "full_name": "Other Restored"},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["is_active"] is True
    assert _login("other@example.com", resp.json()["temp_password"])


def test_third_schedule_csv_upload_prices_and_taxes_off_fixed_value(user_headers):
    # SN008's official sample: fixedNotifiedValueOrRetailPrice=1000, rate=18%.
    # Placed at the end of the file — it adds an extra invoice for the shop
    # user, which would throw off earlier tests' exact invoice-count asserts.
    csv_text = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id,fixed_notified_value\n"
        "POS-3SCH,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Test 3rd Schedule Item,0101.2100,18%,\"Numbers, pieces, units\",1,0,"
        "3rd Schedule Goods,SN008,1000\n"
    )
    resp = client.post(
        "/api/uploads",
        files={"file": ("sn008.csv", csv_text, "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["invoices_submitted"] == 1

    invoices = client.get(
        "/api/invoices?upload_id=" + str(resp.json()["id"]), headers=user_headers
    ).json()
    detail = client.get(f"/api/invoices/{invoices[0]['id']}", headers=user_headers).json()
    item = detail["payload"]["items"][0]
    # FBR's live sandbox rejects a literal 0 here as "invalid" (error 0300)
    # even though it's what PRAL's own spec sample sends — a negligible
    # non-zero placeholder works around that validator quirk.
    assert item["valueSalesExcludingST"] == 0.01
    assert item["fixedNotifiedValueOrRetailPrice"] == 1000
    # Tax must still be computed off the fixed value (1000 × 18% = 180), not
    # the placeholder sale value.
    assert item["salesTaxApplicable"] == 180

    # Our own item serialization (not the FBR payload) also exposes
    # fixed_notified_value, so the receipt can display the real pricing
    # basis instead of the 0.01 placeholder in value_excl_st.
    own_item = detail["items"][0]
    assert own_item["value_excl_st"] == 0.01
    assert own_item["fixed_notified_value"] == 1000


def test_third_schedule_fixed_value_scales_with_quantity(user_headers):
    # fixed_notified_value is entered as a per-unit MRP (what's printed on
    # the package) — selling 4 units must tax 4x the MRP, not 1x. Caught via
    # a real distributor invoice where our receipt showed the same total for
    # a qty-1 and a qty-4 line because quantity was silently dropped from
    # the 3rd Schedule tax basis.
    csv_text = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id,fixed_notified_value\n"
        "POS-3SCH-QTY,2026-08-25,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Test 3rd Schedule Item,0101.2100,18%,\"Numbers, pieces, units\",4,0,"
        "3rd Schedule Goods,SN008,271.19\n"
    )
    resp = client.post(
        "/api/uploads",
        files={"file": ("sn008-qty.csv", csv_text, "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text

    invoices = client.get(
        "/api/invoices?upload_id=" + str(resp.json()["id"]), headers=user_headers
    ).json()
    detail = client.get(f"/api/invoices/{invoices[0]['id']}", headers=user_headers).json()
    item = detail["payload"]["items"][0]
    assert item["fixedNotifiedValueOrRetailPrice"] == 1084.76
    assert item["salesTaxApplicable"] == 195.26


def test_reduced_rate_goods_send_empty_extra_tax_and_tax_off_sale_value(user_headers):
    # "Goods at Reduced Rate" (SN028) has two FBR quirks confirmed live
    # against the real sandbox:
    #  1. extraTax: 0 is rejected as "extra tax provided" (error 0091,
    #     2026-08-17) — must send "" instead, unlike other sale types.
    #  2. Unlike "3rd Schedule Goods" (SN008/SN027), it is NOT taxed off
    #     fixedNotifiedValueOrRetailPrice — FBR rejects salesTaxApplicable
    #     computed that way (2026-08-18: "Provided sales tax amount does
    #     not match the calculated sales tax amount... the provided Sale
    #     Value is used to calculate the Sales Tax"). Tax must be computed
    #     off valueSalesExcludingST instead, matching PRAL's own SN028
    #     sample (salesTaxApplicable: 0, matching valueSalesExcludingST: 0,
    #     even though fixedNotifiedValueOrRetailPrice there is 100).
    csv_text = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id,fixed_notified_value,"
        "sro_schedule_no,sro_item_serial_no\n"
        "POS-REDUCED,2026-08-17,1234567,Test Buyer,Sindh,Karachi,Registered,"
        "Reduced Rate Item,0101.2100,1%,\"Numbers, pieces, units\",1,0,"
        "Goods at Reduced Rate,SN028,100,EIGHTH SCHEDULE Table 1,70\n"
    )
    resp = client.post(
        "/api/uploads",
        files={"file": ("sn028.csv", csv_text, "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["invoices_submitted"] == 1

    invoices = client.get(
        "/api/invoices?upload_id=" + str(resp.json()["id"]), headers=user_headers
    ).json()
    detail = client.get(f"/api/invoices/{invoices[0]['id']}", headers=user_headers).json()
    item = detail["payload"]["items"][0]
    assert item["extraTax"] == ""
    # value_excl (0.01 placeholder, since unit_price=0) must be the tax
    # basis, NOT fixedNotifiedValueOrRetailPrice (100) — 0.01 * 1% rounds
    # to 0.0, matching PRAL's own SN028 sample's salesTaxApplicable of 0.
    assert item["valueSalesExcludingST"] == 0.01
    assert item["fixedNotifiedValueOrRetailPrice"] == 100.0
    assert item["salesTaxApplicable"] == 0.0

    # A standard-rate item (unaffected sale type) must still send numeric 0,
    # matching what's already confirmed working against the live sandbox.
    standard_csv = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id\n"
        "POS-STD,2026-08-17,1234567,Test Buyer,Sindh,Karachi,Registered,"
        "Standard Item,0101.2100,18%,\"Numbers, pieces, units\",1,100,"
        "Goods at standard rate (default),SN001\n"
    )
    resp2 = client.post(
        "/api/uploads",
        files={"file": ("sn001.csv", standard_csv, "text/csv")},
        headers=user_headers,
    )
    invoices2 = client.get(
        "/api/invoices?upload_id=" + str(resp2.json()["id"]), headers=user_headers
    ).json()
    detail2 = client.get(f"/api/invoices/{invoices2[0]['id']}", headers=user_headers).json()
    assert detail2["payload"]["items"][0]["extraTax"] == 0.0


def test_optional_amount_columns_pass_through_to_fbr_payload(user_headers):
    # A provider whose POS already carries a discount / withheld tax / FED
    # etc. can now send those columns and have them reach FBR unchanged,
    # instead of being silently dropped. invoice_ref_no and an explicit
    # sales_tax / total_values override travel through the same way.
    csv_text = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id,invoice_ref_no,"
        "sales_tax,sales_tax_withheld_at_source,extra_tax,further_tax,"
        "fed_payable,discount,total_values\n"
        "POS-CHG,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Discounted Item,0101.2100,18%,\"Numbers, pieces, units\",1,1000,"
        "Goods at standard rate (default),SN002,SI-ORIG-9,180,25,5,40,10,150,1055\n"
    )
    resp = client.post(
        "/api/uploads",
        files={"file": ("charges.csv", csv_text, "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["invoices_submitted"] == 1

    invoices = client.get(
        "/api/invoices?upload_id=" + str(resp.json()["id"]), headers=user_headers
    ).json()
    detail = client.get(f"/api/invoices/{invoices[0]['id']}", headers=user_headers).json()
    payload = detail["payload"]
    assert payload["invoiceRefNo"] == "SI-ORIG-9"
    item = payload["items"][0]
    assert item["salesTaxApplicable"] == 180
    assert item["salesTaxWithheldAtSource"] == 25
    assert item["extraTax"] == 5
    assert item["furtherTax"] == 40
    assert item["fedPayable"] == 10
    assert item["discount"] == 150
    # Explicit total_values wins over the derived line total.
    assert item["totalValues"] == 1055

    # An old-style file with none of these columns still works, and the
    # amounts default to 0 / the derived total.
    plain = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id\n"
        "POS-PLAIN,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Plain Item,0101.2100,18%,\"Numbers, pieces, units\",1,1000,"
        "Goods at standard rate (default),SN002\n"
    )
    resp2 = client.post(
        "/api/uploads",
        files={"file": ("plain.csv", plain, "text/csv")},
        headers=user_headers,
    )
    inv2 = client.get(
        "/api/invoices?upload_id=" + str(resp2.json()["id"]), headers=user_headers
    ).json()
    d2 = client.get(f"/api/invoices/{inv2[0]['id']}", headers=user_headers).json()
    it2 = d2["payload"]["items"][0]
    assert it2["discount"] == 0
    assert it2["furtherTax"] == 0
    assert it2["totalValues"] == 1180  # 1000 + 180 sales tax, nothing else


def test_csv_template_carries_optional_amount_columns(user_headers):
    resp = client.get("/api/uploads/template", headers=user_headers)
    assert resp.status_code == 200
    header = resp.text.splitlines()[0].split(",")
    for col in (
        "invoice_ref_no",
        "sales_tax",
        "sales_tax_withheld_at_source",
        "extra_tax",
        "further_tax",
        "fed_payable",
        "discount",
        "total_values",
        "advance_tax",
    ):
        assert col in header


def test_advance_tax_from_csv_is_summed_and_locked(user_headers):
    # advance_tax rows sharing a pos_invoice_no add up onto the invoice; it's
    # a receipt figure (not an item field, so not in the payload), and an
    # upload invoice is "set" so the receipt back-fill is not offered.
    csv_text = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id,advance_tax\n"
        "POS-AIT,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Item A,0101.2100,18%,\"Numbers, pieces, units\",1,1000,"
        "Goods at standard rate (default),SN002,6.25\n"
        "POS-AIT,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "Item B,0101.2100,18%,\"Numbers, pieces, units\",1,500,"
        "Goods at standard rate (default),SN002,3.75\n"
    )
    up = client.post(
        "/api/uploads",
        files={"file": ("ait.csv", csv_text, "text/csv")},
        headers=user_headers,
    ).json()
    inv = client.get(
        f"/api/invoices?upload_id={up['id']}", headers=user_headers
    ).json()[0]
    detail = client.get(f"/api/invoices/{inv['id']}", headers=user_headers).json()
    assert detail["advance_tax"] == 10.0          # 6.25 + 3.75
    assert detail["advance_tax_set"] is True
    assert "advanceTax" not in detail["payload"]
    assert "advance_tax" not in str(detail["payload"]).lower()

    # Already set -> the one-time back-fill is refused.
    resp = client.patch(
        f"/api/invoices/{inv['id']}/advance-tax",
        json={"advance_tax": 99},
        headers=user_headers,
    )
    assert resp.status_code == 400


def test_advance_tax_backfill_sets_once(user_headers):
    from app.database import SessionLocal
    from app.models import Invoice

    up = client.post(
        "/api/uploads",
        files={"file": ("bf.csv", CSV_OK, "text/csv")},
        headers=user_headers,
    ).json()
    inv_id = client.get(
        f"/api/invoices?upload_id={up['id']}", headers=user_headers
    ).json()[0]["id"]

    # Simulate an older, pre-feature invoice.
    with SessionLocal() as db:
        db.get(Invoice, inv_id).advance_tax_set = False
        db.commit()

    d0 = client.get(f"/api/invoices/{inv_id}", headers=user_headers).json()
    assert d0["advance_tax"] == 0.0 and d0["advance_tax_set"] is False

    assert (
        client.patch(
            f"/api/invoices/{inv_id}/advance-tax",
            json={"advance_tax": -1},
            headers=user_headers,
        ).status_code
        == 400
    )

    ok = client.patch(
        f"/api/invoices/{inv_id}/advance-tax",
        json={"advance_tax": 123.456},
        headers=user_headers,
    )
    assert ok.status_code == 200, ok.text
    assert ok.json()["advance_tax"] == 123.46 and ok.json()["advance_tax_set"] is True

    # Locked now.
    assert (
        client.patch(
            f"/api/invoices/{inv_id}/advance-tax",
            json={"advance_tax": 0},
            headers=user_headers,
        ).status_code
        == 400
    )


def test_compute_sales_tax_handles_percent_and_fixed_per_unit_rates():
    from app.services.invoice_service import compute_sales_tax

    # Plain percentage (unchanged behaviour).
    assert compute_sales_tax(1000, "18%") == 180.0
    assert compute_sales_tax(100, "1.43%") == 1.43
    assert compute_sales_tax(1000, "0%") == 0.0
    assert compute_sales_tax(1000, "Exempt") == 0.0
    # Fixed rupees per unit x quantity — PRAL SN021 ("Rs.3", qty 12 -> 36)
    # and SN023 ("Rs.200", qty 123 -> 24600).
    assert compute_sales_tax(123, "Rs.3", 12) == 36.0
    assert compute_sales_tax(234, "Rs.200", 123) == 24600.0
    # Percentage + per-unit together — PRAL SN022 ("18% along with rupees
    # 60 per kilogram", value 100, qty 1 -> 18 + 60 = 78).
    assert compute_sales_tax(100, "18% along with rupees 60 per kilogram", 1) == 78.0


def test_csv_upload_with_fixed_per_unit_rate(user_headers):
    # SN023-style: rate is a flat "Rs.200" per unit, so tax must be
    # 200 x quantity, not 0 (which is what a "%"-only parser produced).
    csv_text = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id\n"
        "POS-CNG,2026-08-17,,Walk-in Customer,Sindh,Karachi,Unregistered,"
        "CNG,0101.2100,Rs.200,\"Numbers, pieces, units\",123,234,CNG Sales,SN023\n"
    )
    resp = client.post(
        "/api/uploads",
        files={"file": ("sn023.csv", csv_text, "text/csv")},
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    invoices = client.get(
        "/api/invoices?upload_id=" + str(resp.json()["id"]), headers=user_headers
    ).json()
    detail = client.get(f"/api/invoices/{invoices[0]['id']}", headers=user_headers).json()
    assert detail["payload"]["items"][0]["salesTaxApplicable"] == 24600.0


def test_admin_views_user_uploads_and_invoices(admin_headers, user_headers):
    users = client.get("/api/admin/users", headers=admin_headers).json()
    shop = next(u for u in users if u["email"] == "shop@example.com")

    own_uploads = client.get("/api/uploads", headers=user_headers).json()
    admin_view_uploads = client.get(
        f"/api/admin/users/{shop['id']}/uploads", headers=admin_headers
    ).json()
    assert admin_view_uploads == own_uploads
    assert len(admin_view_uploads) > 0

    own_invoices = client.get("/api/invoices", headers=user_headers).json()
    admin_view_invoices = client.get(
        f"/api/admin/users/{shop['id']}/invoices", headers=admin_headers
    ).json()
    assert admin_view_invoices == own_invoices
    assert len(admin_view_invoices) > 0

    inv_id = own_invoices[0]["id"]
    own_detail = client.get(f"/api/invoices/{inv_id}", headers=user_headers).json()
    admin_detail = client.get(
        f"/api/admin/users/{shop['id']}/invoices/{inv_id}", headers=admin_headers
    ).json()
    assert admin_detail == own_detail

    # upload_id filter behaves the same as the user's own endpoint.
    upload_id = own_uploads[0]["id"]
    filtered_own = client.get(
        f"/api/invoices?upload_id={upload_id}", headers=user_headers
    ).json()
    filtered_admin = client.get(
        f"/api/admin/users/{shop['id']}/invoices?upload_id={upload_id}",
        headers=admin_headers,
    ).json()
    assert filtered_admin == filtered_own


def test_admin_views_reject_nonexistent_user(admin_headers):
    assert client.get("/api/admin/users/999999/uploads", headers=admin_headers).status_code == 404
    assert client.get("/api/admin/users/999999/invoices", headers=admin_headers).status_code == 404
    assert client.get("/api/admin/users/999999/invoices/1", headers=admin_headers).status_code == 404


def test_admin_invoice_detail_checks_ownership(admin_headers, user_headers):
    # An invoice ID that belongs to a different user must 404 under the
    # wrong user_id in the path — not just "any invoice by ID".
    users = client.get("/api/admin/users", headers=admin_headers).json()
    admin_self = next(u for u in users if u["email"] == "admin@example.com")
    invoices = client.get("/api/invoices", headers=user_headers).json()
    resp = client.get(
        f"/api/admin/users/{admin_self['id']}/invoices/{invoices[0]['id']}",
        headers=admin_headers,
    )
    assert resp.status_code == 404


def test_non_admin_cannot_view_admin_user_views(user_headers):
    assert client.get("/api/admin/users/1/uploads", headers=user_headers).status_code == 403
    assert client.get("/api/admin/users/1/invoices", headers=user_headers).status_code == 403


def test_admin_stats_extended_fields(admin_headers):
    stats = client.get("/api/admin/stats", headers=admin_headers).json()
    assert stats["total_admins"] >= 1
    assert stats["total_regular_users"] >= 1
    assert stats["total_admins"] + stats["total_regular_users"] == stats["total_users"]
    assert "draft_invoices" in stats
    assert len(stats["invoices_by_day"]) == 14
    for bucket in stats["invoices_by_day"]:
        assert set(bucket) == {"date", "total", "submitted", "failed"}
    # Invoices created earlier in this test run must show up in today's bucket.
    assert stats["invoices_by_day"][-1]["total"] >= 1


def test_create_admin_via_admin_users_endpoint(admin_headers):
    resp = client.post(
        "/api/admin/users",
        json={"email": "second-admin@example.com", "full_name": "Second Admin", "role": "admin"},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["role"] == "admin"
    users = client.get("/api/admin/users", headers=admin_headers).json()
    created = next(u for u in users if u["email"] == "second-admin@example.com")
    assert created["role"] == "admin"


def test_create_user_gets_auto_generated_6_digit_temp_password(admin_headers):
    resp = client.post(
        "/api/admin/users",
        json={"email": "autogen@example.com", "full_name": "Auto Gen"},
        headers=admin_headers,
    )
    assert resp.status_code == 201, resp.text
    temp_password = resp.json()["temp_password"]
    assert len(temp_password) == 6
    assert temp_password.isalnum() and temp_password == temp_password.upper()
    # It's a real, working credential for the one first login.
    login = client.post(
        "/api/auth/login", json={"email": "autogen@example.com", "password": temp_password}
    )
    assert login.status_code == 200
    assert login.json()["must_change_password"] is True

    # A client-supplied password is ignored — the server always generates
    # its own, so there's no way to end up with a weak/guessable temp code.
    resp2 = client.post(
        "/api/admin/users",
        json={
            "email": "autogen2@example.com",
            "full_name": "Auto Gen 2",
            "password": "ignored-if-sent",
        },
        headers=admin_headers,
    )
    assert resp2.status_code == 201, resp2.text
    temp2 = resp2.json()["temp_password"]
    assert temp2 != "ignored-if-sent"
    assert len(temp2) == 6 and temp2.isalnum() and temp2 == temp2.upper()


def test_users_list_role_filter(admin_headers):
    resp = client.get("/api/admin/users?role=admin", headers=admin_headers)
    assert resp.status_code == 200
    assert all(u["role"] == "admin" for u in resp.json())
    assert len(resp.json()) >= 1

    resp = client.get("/api/admin/users?role=user", headers=admin_headers)
    assert all(u["role"] == "user" for u in resp.json())

    assert client.get("/api/admin/users?role=bogus", headers=admin_headers).status_code == 400


def test_users_list_search_filter(admin_headers):
    resp = client.get("/api/admin/users?q=shop@example.com", headers=admin_headers)
    emails = [u["email"] for u in resp.json()]
    assert emails == ["shop@example.com"]

    resp = client.get("/api/admin/users?q=Shop Owner", headers=admin_headers)
    assert "shop@example.com" in [u["email"] for u in resp.json()]

    resp = client.get("/api/admin/users?q=no-such-person-xyz", headers=admin_headers)
    assert resp.json() == []


def test_users_list_active_filter(admin_headers):
    resp = client.get("/api/admin/users?is_active=true", headers=admin_headers)
    assert all(u["is_active"] for u in resp.json())
    resp = client.get("/api/admin/users?is_active=false", headers=admin_headers)
    assert all(not u["is_active"] for u in resp.json())


def test_users_list_pagination(admin_headers):
    full = client.get("/api/admin/users", headers=admin_headers)
    total = int(full.headers["x-total-count"])
    all_users = full.json()
    assert total == len(all_users)
    assert total >= 3  # admin + shop owner + second-admin created above

    resp = client.get("/api/admin/users?page=1&page_size=2", headers=admin_headers)
    assert int(resp.headers["x-total-count"]) == total
    page1 = resp.json()
    assert len(page1) == 2
    assert page1 == all_users[0:2]  # same default order (created_at desc)

    resp2 = client.get("/api/admin/users?page=2&page_size=2", headers=admin_headers)
    page2 = resp2.json()
    assert page2 == all_users[2:4]
    # No overlap between pages.
    assert {u["id"] for u in page1}.isdisjoint({u["id"] for u in page2})

    assert client.get("/api/admin/users?page=0", headers=admin_headers).status_code == 400
    assert client.get("/api/admin/users?page_size=0", headers=admin_headers).status_code == 400
    assert client.get("/api/admin/users?page_size=1001", headers=admin_headers).status_code == 400


def test_uploads_list_pagination_and_filters(admin_headers, user_headers):
    full = client.get("/api/uploads", headers=user_headers)
    total = int(full.headers["x-total-count"])
    all_uploads = full.json()
    assert total == len(all_uploads)
    assert total >= 4

    resp = client.get("/api/uploads?page=1&page_size=2", headers=user_headers)
    assert int(resp.headers["x-total-count"]) == total
    page1 = resp.json()
    assert len(page1) == 2
    assert page1 == all_uploads[0:2]  # same default order (id desc)

    resp2 = client.get("/api/uploads?page=2&page_size=2", headers=user_headers)
    assert resp2.json() == all_uploads[2:4]

    # Create our own guaranteed-failed upload rather than depending on one
    # an earlier test happened to leave behind.
    client.post(
        "/api/uploads",
        files={"file": ("malformed.csv", "foo,bar\n1,2\n", "text/csv")},
        headers=user_headers,
    )

    failed = client.get("/api/uploads?status=failed", headers=user_headers).json()
    assert len(failed) >= 1
    assert all(u["status"] == "failed" for u in failed)
    assert any(u["filename"] == "malformed.csv" for u in failed)

    named = client.get("/api/uploads?q=malformed.csv", headers=user_headers).json()
    assert len(named) == 1

    assert client.get("/api/uploads?status=bogus", headers=user_headers).status_code == 400
    assert client.get("/api/uploads?page_size=0", headers=user_headers).status_code == 400

    # The admin read-only mirror returns byte-identical data, same filters.
    users = client.get("/api/admin/users", headers=admin_headers).json()
    shop = next(u for u in users if u["email"] == "shop@example.com")
    admin_full = client.get(f"/api/admin/users/{shop['id']}/uploads", headers=admin_headers)
    assert int(admin_full.headers["x-total-count"]) == total + 1
    admin_failed = client.get(
        f"/api/admin/users/{shop['id']}/uploads?status=failed", headers=admin_headers
    ).json()
    assert admin_failed == failed


def test_invoices_list_pagination_and_filters(admin_headers, user_headers):
    full = client.get("/api/invoices", headers=user_headers)
    total = int(full.headers["x-total-count"])
    all_invoices = full.json()
    assert total == len(all_invoices)
    assert total >= 5

    resp = client.get("/api/invoices?page=2&page_size=3", headers=user_headers)
    assert int(resp.headers["x-total-count"]) == total
    assert resp.json() == all_invoices[3:6]

    # Manufacture a failed invoice (quantity 0 trips the mock validator) so
    # the status filter has something real to find.
    bad_csv = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id\n"
        "POS-BADQTY,2026-08-17,1234567,Test Buyer,Sindh,Karachi,Registered,"
        "Zero Qty Item,0101.2100,18%,\"Numbers, pieces, units\",0,100,"
        "Goods at standard rate (default),SN001\n"
    )
    upload_resp = client.post(
        "/api/uploads",
        files={"file": ("badqty.csv", bad_csv, "text/csv")},
        headers=user_headers,
    )
    assert upload_resp.json()["invoices_failed"] == 1

    failed = client.get("/api/invoices?status=failed", headers=user_headers).json()
    assert len(failed) == 1
    assert failed[0]["pos_invoice_no"] == "POS-BADQTY"

    by_buyer = client.get("/api/invoices?q=ABC Traders", headers=user_headers).json()
    assert len(by_buyer) >= 1
    assert all(i["buyer_name"] == "ABC Traders" for i in by_buyer)

    by_pos_no = client.get("/api/invoices?q=POS-BADQTY", headers=user_headers).json()
    assert len(by_pos_no) == 1

    assert client.get("/api/invoices?status=bogus", headers=user_headers).status_code == 400

    # Admin mirror matches exactly, including the status filter.
    users = client.get("/api/admin/users", headers=admin_headers).json()
    shop = next(u for u in users if u["email"] == "shop@example.com")
    admin_failed = client.get(
        f"/api/admin/users/{shop['id']}/invoices?status=failed", headers=admin_headers
    ).json()
    assert admin_failed == failed


def test_my_stats_scoped_to_current_user(admin_headers, user_headers):
    own_uploads = client.get("/api/uploads", headers=user_headers)
    own_uploads_total = int(own_uploads.headers["x-total-count"])
    own_invoices = client.get("/api/invoices", headers=user_headers)
    own_invoices_total = int(own_invoices.headers["x-total-count"])
    submitted = [i for i in own_invoices.json() if i["status"] == "submitted"]
    failed = [i for i in own_invoices.json() if i["status"] == "failed"]

    stats = client.get("/api/stats", headers=user_headers).json()
    assert stats["total_uploads"] == own_uploads_total
    assert stats["total_invoices"] == own_invoices_total
    assert stats["submitted_invoices"] == len(submitted)
    assert stats["failed_invoices"] == len(failed)
    assert stats["total_sales_value"] == round(sum(i["grand_total"] for i in submitted), 2)
    assert stats["total_tax_collected"] == round(sum(i["total_tax"] for i in submitted), 2)
    assert len(stats["invoices_by_day"]) == 14
    for bucket in stats["invoices_by_day"]:
        assert set(bucket) == {"date", "total", "submitted", "failed"}
    assert stats["invoices_by_day"][-1]["total"] >= 1

    # A second, brand-new user with zero activity sees all zeros — not the
    # first user's data.
    resp = client.post(
        "/api/admin/users",
        json={"email": "quiet@example.com", "full_name": "Quiet One"},
        headers=admin_headers,
    )
    login = client.post(
        "/api/auth/login",
        json={"email": "quiet@example.com", "password": resp.json()["temp_password"]},
    )
    quiet_headers = {"Authorization": f"Bearer {login.json()['token']}"}
    quiet_resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "Str0ng!Passw0rd99", "confirm_password": "Str0ng!Passw0rd99"},
        headers=quiet_headers,
    )
    assert quiet_resp.status_code == 200
    quiet_headers = {
        "Authorization": "Bearer "
        + client.post(
            "/api/auth/login",
            json={"email": "quiet@example.com", "password": "Str0ng!Passw0rd99"},
        ).json()["token"]
    }
    quiet_stats = client.get("/api/stats", headers=quiet_headers).json()
    assert quiet_stats["total_uploads"] == 0
    assert quiet_stats["total_invoices"] == 0
    assert quiet_stats["submitted_invoices"] == 0
    assert quiet_stats["total_sales_value"] == 0
    assert all(b["total"] == 0 for b in quiet_stats["invoices_by_day"])

    assert client.get("/api/stats").status_code == 401


def _one_invoice(headers, csv_text, target):
    up = client.post(
        "/api/uploads",
        files={"file": (f"{target}.csv", csv_text, "text/csv")},
        data={"target": target},
        headers=headers,
    ).json()
    return client.get(f"/api/invoices?upload_id={up['id']}", headers=headers).json()[0]


def test_mark_invoice_paid_requires_submitted_live_invoice(admin_headers):
    headers, _ = _make_account(
        admin_headers, "paidguard@example.com", can_submit_production=True
    )

    # A live (production) submitted invoice: mark paid, toggle back.
    live = _one_invoice(headers, _ENV_CSV, "production")
    assert live["fbr_env"] == "production" and live["is_paid"] is False
    assert (
        client.patch(
            f"/api/invoices/{live['id']}/paid", json={"is_paid": True}, headers=headers
        ).json()["is_paid"]
        is True
    )
    assert (
        client.patch(
            f"/api/invoices/{live['id']}/paid", json={"is_paid": False}, headers=headers
        ).json()["is_paid"]
        is False
    )

    # A test (sandbox) invoice — submitted, but not a real record → 400.
    test_inv = _one_invoice(headers, _ENV_CSV, "sandbox")
    assert test_inv["fbr_env"] == "sandbox" and test_inv["status"] == "submitted"
    assert (
        client.patch(
            f"/api/invoices/{test_inv['id']}/paid", json={"is_paid": True}, headers=headers
        ).status_code
        == 400
    )

    # A failed invoice was never issued → 400.
    bad_csv = _ENV_CSV.replace(
        "\"Numbers, pieces, units\",1,1000", "\"Numbers, pieces, units\",0,1000"
    )
    bad = _one_invoice(headers, bad_csv, "production")
    assert bad["status"] == "failed"
    assert (
        client.patch(
            f"/api/invoices/{bad['id']}/paid", json={"is_paid": True}, headers=headers
        ).status_code
        == 400
    )


def test_user_can_delete_test_invoice_only(admin_headers):
    headers, _ = _make_account(
        admin_headers, "invdel@example.com", can_submit_production=True
    )

    # A test (sandbox) invoice — the user can remove it from their history.
    test_inv = _one_invoice(headers, _ENV_CSV, "sandbox")
    assert test_inv["fbr_env"] == "sandbox"
    assert client.delete(
        f"/api/invoices/{test_inv['id']}", headers=headers
    ).json() == {"ok": True}
    remaining = client.get("/api/invoices", headers=headers).json()
    assert all(i["id"] != test_inv["id"] for i in remaining)
    # Gone from the user's view, and gone for the admin too.
    assert client.get(f"/api/invoices/{test_inv['id']}", headers=headers).status_code == 404

    # A live (production) invoice is a real record — 403, still there.
    live = _one_invoice(headers, _ENV_CSV, "production")
    assert live["fbr_env"] == "production"
    assert (
        client.delete(f"/api/invoices/{live['id']}", headers=headers).status_code == 403
    )
    still = client.get("/api/invoices?fbr_env=production", headers=headers).json()
    assert any(i["id"] == live["id"] for i in still)

    # Another user can't touch it.
    other, _ = _make_account(admin_headers, "invdel-other@example.com")
    assert (
        client.delete(f"/api/invoices/{live['id']}", headers=other).status_code == 404
    )


def test_paid_tax_reflected_in_stats(admin_headers):
    headers, _ = _make_account(
        admin_headers, "paidstats@example.com", can_submit_production=True
    )
    inv = _one_invoice(headers, _ENV_CSV, "production")
    client.patch(f"/api/invoices/{inv['id']}/paid", json={"is_paid": True}, headers=headers)

    stats = client.get("/api/stats", headers=headers).json()
    assert stats["paid_tax"] == inv["total_tax"]

    admin_stats = client.get("/api/admin/stats", headers=admin_headers).json()
    assert admin_stats["paid_tax"] >= inv["total_tax"]
    assert admin_stats["total_tax_collected"] >= admin_stats["paid_tax"]


def test_admin_user_growth_chart(admin_headers):
    for granularity, expected_periods in [
        ("day", 30),
        ("week", 12),
        ("month", 12),
        ("year", 5),
    ]:
        resp = client.get(
            f"/api/admin/stats/user-growth?granularity={granularity}", headers=admin_headers
        )
        assert resp.status_code == 200, resp.text
        buckets = resp.json()
        assert len(buckets) == expected_periods
        for b in buckets:
            assert set(b) == {"period", "label", "count"}
        # Every account created during this test run must land in some
        # bucket — total counts across all buckets can't be zero.
        assert sum(b["count"] for b in buckets) >= 1

    resp = client.get("/api/admin/stats/user-growth?granularity=bogus", headers=admin_headers)
    assert resp.status_code == 400


def test_excel_upload_happy_path(user_headers):
    import io

    import openpyxl

    from app.services.csv_processor import ALL_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ALL_COLUMNS)
    row = [
        "POS-XL1",
        "2026-08-19",
        "1234567",
        "Excel Buyer",
        "Punjab",
        "Lahore",
        "Registered",
        "Excel Item",
        "8471.3010",
        "18%",
        "Numbers, pieces, units",
        3,
        1000,
        "Goods at standard rate (default)",
        "SN001",
    ]
    ws.append(row + [""] * (len(ALL_COLUMNS) - len(row)))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/uploads",
        files={
            "file": (
                "sales.xlsx",
                buf.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    upload = resp.json()
    assert upload["status"] == "completed"
    assert upload["total_rows"] == 1
    assert upload["invoices_created"] == 1
    assert upload["invoices_submitted"] == 1

    invoices = client.get(
        f"/api/invoices?upload_id={upload['id']}", headers=user_headers
    ).json()
    assert invoices[0]["total_excl"] == 3000.0
    assert invoices[0]["total_tax"] == 540.0


def test_excel_upload_with_percent_and_date_formatted_cells(user_headers):
    # Regression: typing "18%" into an Excel cell stores the underlying
    # value as 0.18 with a percentage number format, and a date-typed cell
    # comes back as a Python datetime — both must be normalized back to
    # the plain strings our validation/pricing logic expects.
    import io

    import openpyxl

    from app.services.csv_processor import ALL_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ALL_COLUMNS)
    row_idx = 2
    row = [
        "POS-XL2",
        date(2026, 8, 19),
        "",
        "Walk-in Customer",
        "Sindh",
        "Karachi",
        "Unregistered",
        "Formatted Item",
        "8517.1219",
        0.18,
        "Numbers, pieces, units",
        1,
        1000,
        "Goods at standard rate (default)",
        "SN001",
    ]
    ws.append(row + [""] * (len(ALL_COLUMNS) - len(row)))
    rate_cell = ws.cell(row=row_idx, column=ALL_COLUMNS.index("rate") + 1)
    rate_cell.number_format = "0%"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/uploads",
        files={
            "file": (
                "formatted.xlsx",
                buf.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    upload = resp.json()
    assert upload["status"] == "completed", upload
    assert upload["invoices_submitted"] == 1

    invoices = client.get(
        f"/api/invoices?upload_id={upload['id']}", headers=user_headers
    ).json()
    # 1000 * 18% = 180, proving the rate cell was read back as "18%" (not
    # silently treated as a 0.18% rate).
    assert invoices[0]["total_tax"] == 180.0


def test_excel_upload_corrupt_file_gives_clear_error(user_headers):
    resp = client.post(
        "/api/uploads",
        files={
            "file": (
                "not-really-excel.xlsx",
                b"this is not a real xlsx file",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["status"] == "failed"
    assert "Excel file" in body["error"]


def test_upload_rejects_unsupported_extension(user_headers):
    resp = client.post(
        "/api/uploads",
        files={"file": ("sales.txt", "pos_invoice_no,invoice_date\n", "text/plain")},
        headers=user_headers,
    )
    assert resp.status_code == 400
    assert ".csv or .xlsx" in resp.json()["detail"]


def test_expired_token_rejected_with_clear_message(user_headers):
    settings = get_settings()
    expired_token = jwt.encode(
        {
            "sub": "1",
            "role": "user",
            "tv": 0,
            "exp": datetime.now(timezone.utc) - timedelta(hours=1),
        },
        settings.jwt_secret,
        algorithm="HS256",
    )
    resp = client.get(
        "/api/uploads", headers={"Authorization": f"Bearer {expired_token}"}
    )
    assert resp.status_code == 401
    assert "expired" in resp.json()["detail"].lower()


def test_change_password_flow(admin_headers):
    # A dedicated throwaway account — must not touch the shared user_headers
    # fixture's token, since change-password invalidates it.
    resp = client.post(
        "/api/admin/users",
        json={"email": "changepw@example.com", "full_name": "Change PW"},
        headers=admin_headers,
    )
    temp_password = resp.json()["temp_password"]
    login = client.post(
        "/api/auth/login", json={"email": "changepw@example.com", "password": temp_password}
    )
    temp_headers = {"Authorization": f"Bearer {login.json()['token']}"}
    set_resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "FirstReal8", "confirm_password": "FirstReal8"},
        headers=temp_headers,
    )
    assert set_resp.status_code == 200, set_resp.text
    headers = {"Authorization": f"Bearer {set_resp.json()['token']}"}

    # Wrong current password rejected — 400, not 401: the bearer token is
    # perfectly valid here, so this must not trigger the frontend's
    # expired-session auto-logout (which fires on any 401 with a token).
    resp = client.post(
        "/api/auth/change-password",
        json={
            "current_password": "WrongPass1",
            "new_password": "SecondReal8",
            "confirm_password": "SecondReal8",
        },
        headers=headers,
    )
    assert resp.status_code == 400
    assert client.get("/api/auth/me", headers=headers).status_code == 200  # token still valid

    # Mismatched confirmation rejected.
    resp = client.post(
        "/api/auth/change-password",
        json={
            "current_password": "FirstReal8",
            "new_password": "SecondReal8",
            "confirm_password": "Different1",
        },
        headers=headers,
    )
    assert resp.status_code == 400

    # Too-short new password rejected (the only real rule now: 8+ chars).
    resp = client.post(
        "/api/auth/change-password",
        json={
            "current_password": "FirstReal8",
            "new_password": "short1",
            "confirm_password": "short1",
        },
        headers=headers,
    )
    assert resp.status_code == 400

    # Success: old token invalidated, new one works, new password logs in.
    resp = client.post(
        "/api/auth/change-password",
        json={
            "current_password": "FirstReal8",
            "new_password": "SecondReal8",
            "confirm_password": "SecondReal8",
        },
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    fresh_headers = {"Authorization": f"Bearer {resp.json()['token']}"}

    assert client.get("/api/auth/me", headers=headers).status_code == 401
    assert client.get("/api/auth/me", headers=fresh_headers).status_code == 200

    assert (
        client.post(
            "/api/auth/login",
            json={"email": "changepw@example.com", "password": "SecondReal8"},
        ).status_code
        == 200
    )
    assert (
        client.post(
            "/api/auth/login",
            json={"email": "changepw@example.com", "password": "FirstReal8"},
        ).status_code
        == 401
    )


def test_admin_reset_password(admin_headers):
    resp = client.post(
        "/api/admin/users",
        json={"email": "resetpw@example.com", "full_name": "Reset PW"},
        headers=admin_headers,
    )
    user_id = resp.json()["id"]
    original_temp = resp.json()["temp_password"]
    login = client.post(
        "/api/auth/login", json={"email": "resetpw@example.com", "password": original_temp}
    )
    old_headers = {"Authorization": f"Bearer {login.json()['token']}"}
    set_resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "RealPassw0rd", "confirm_password": "RealPassw0rd"},
        headers=old_headers,
    )
    real_headers = {"Authorization": f"Bearer {set_resp.json()['token']}"}
    assert client.get("/api/auth/me", headers=real_headers).status_code == 200

    reset_resp = client.post(
        f"/api/admin/users/{user_id}/reset-password", headers=admin_headers
    )
    assert reset_resp.status_code == 200, reset_resp.text
    new_temp = reset_resp.json()["temp_password"]
    assert len(new_temp) == 6
    assert new_temp.isalnum() and new_temp == new_temp.upper()
    assert reset_resp.json()["must_change_password"] is True

    # The old session is dead immediately.
    assert client.get("/api/auth/me", headers=real_headers).status_code == 401

    # Old password no longer works; the new temp code does, and forces
    # must_change_password again.
    assert (
        client.post(
            "/api/auth/login",
            json={"email": "resetpw@example.com", "password": "RealPassw0rd"},
        ).status_code
        == 401
    )
    relogin = client.post(
        "/api/auth/login", json={"email": "resetpw@example.com", "password": new_temp}
    )
    assert relogin.status_code == 200
    assert relogin.json()["must_change_password"] is True


def test_admin_cannot_reset_own_password(admin_headers):
    me = client.get("/api/auth/me", headers=admin_headers).json()
    resp = client.post(f"/api/admin/users/{me['id']}/reset-password", headers=admin_headers)
    assert resp.status_code == 400


def test_non_admin_cannot_reset_password(admin_headers, user_headers):
    users = client.get("/api/admin/users", headers=admin_headers).json()
    shop = next(u for u in users if u["email"] == "shop@example.com")
    resp = client.post(f"/api/admin/users/{shop['id']}/reset-password", headers=user_headers)
    assert resp.status_code == 403


def test_admin_deletes_upload_cascades_all_its_invoices(admin_headers, user_headers):
    # A dedicated throwaway account so this doesn't disturb invoice counts
    # other tests assert on via the shared shop@example.com user.
    resp = client.post(
        "/api/admin/users",
        json={"email": "hscascade@example.com", "full_name": "HS Cascade"},
        headers=admin_headers,
    )
    user_id = resp.json()["id"]
    temp_password = resp.json()["temp_password"]
    client.put(
        f"/api/admin/users/{user_id}/fbr-settings",
        json={
            "fbr_env": "mock",
            "seller_ntn_cnic": "1112223",
            "seller_business_name": "HS Cascade Pvt Ltd",
            "seller_province": "Punjab",
            "seller_address": "Lahore",
            "default_scenario": "SN001",
        },
        headers=admin_headers,
    )
    login = client.post(
        "/api/auth/login", json={"email": "hscascade@example.com", "password": temp_password}
    )
    temp_headers = {"Authorization": f"Bearer {login.json()['token']}"}
    set_resp = client.post(
        "/api/auth/set-password",
        json={"new_password": "CascadeReal8", "confirm_password": "CascadeReal8"},
        headers=temp_headers,
    )
    cascade_headers = {"Authorization": f"Bearer {set_resp.json()['token']}"}

    # One good row (→ submitted) and one quantity-0 row (→ trips the mock
    # validator and fails) in the same upload.
    mixed_csv = (
        "pos_invoice_no,invoice_date,buyer_ntn_cnic,buyer_name,buyer_province,"
        "buyer_address,buyer_registration_type,product_description,hs_code,"
        "rate,uom,quantity,unit_price,sale_type,scenario_id\n"
        "POS-GOOD,2026-08-17,1234567,Good Buyer,Punjab,Lahore,Registered,"
        "Good Item,0101.2100,18%,\"Numbers, pieces, units\",1,1000,"
        "Goods at standard rate (default),SN001\n"
        "POS-BAD,2026-08-17,1234567,Bad Buyer,Punjab,Lahore,Registered,"
        "Zero Qty Item,0101.2100,18%,\"Numbers, pieces, units\",0,100,"
        "Goods at standard rate (default),SN001\n"
    )
    upload_resp = client.post(
        "/api/uploads",
        files={"file": ("mixed.csv", mixed_csv, "text/csv")},
        headers=cascade_headers,
    )
    assert upload_resp.status_code == 201, upload_resp.text
    upload = upload_resp.json()
    assert upload["invoices_submitted"] == 1
    assert upload["invoices_failed"] == 1

    invoices_before = client.get(
        f"/api/admin/users/{user_id}/invoices", headers=admin_headers
    ).json()
    assert len(invoices_before) == 2
    submitted_inv = next(i for i in invoices_before if i["status"] == "submitted")
    failed_inv = next(i for i in invoices_before if i["status"] == "failed")

    del_resp = client.delete(
        f"/api/admin/users/{user_id}/uploads/{upload['id']}", headers=admin_headers
    )
    assert del_resp.status_code == 200, del_resp.text
    assert del_resp.json() == {"ok": True, "invoices_hidden": 2}

    # Upload itself is gone from history.
    remaining_uploads = client.get(
        f"/api/admin/users/{user_id}/uploads", headers=admin_headers
    ).json()
    assert all(u["id"] != upload["id"] for u in remaining_uploads)

    # The whole batch is hidden — the failed row and the submitted one alike.
    remaining_invoices = client.get(
        f"/api/admin/users/{user_id}/invoices", headers=admin_headers
    ).json()
    assert remaining_invoices == []
    assert client.get(
        f"/api/admin/users/{user_id}/invoices/{failed_inv['id']}", headers=admin_headers
    ).status_code == 404
    assert client.get(
        f"/api/admin/users/{user_id}/invoices/{submitted_inv['id']}", headers=admin_headers
    ).status_code == 404

    own_invoices = client.get("/api/invoices", headers=cascade_headers).json()
    assert own_invoices == []

    # Already-deleted upload (and a nonexistent one) both 404 on retry.
    assert (
        client.delete(
            f"/api/admin/users/{user_id}/uploads/{upload['id']}", headers=admin_headers
        ).status_code
        == 404
    )
    assert (
        client.delete(
            f"/api/admin/users/{user_id}/uploads/999999", headers=admin_headers
        ).status_code
        == 404
    )

    # Non-admin can't call this at all.
    assert (
        client.delete(
            f"/api/admin/users/{user_id}/uploads/{upload['id']}", headers=user_headers
        ).status_code
        == 403
    )
